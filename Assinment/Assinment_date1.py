import calendar
import time
import xlsxwriter
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook

# C:\Users\Farhan\PycharmProjects\Selenium_python
my_date = date.today()

workbook = Workbook()

wb = load_workbook("C:\\Users\\Farhan\\PycharmProjects\\Selenium_python\\folder\\date.xlsx")
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://www.google.com")
wait = WebDriverWait(driver, 200)
ws_week = wb[calendar.day_name[my_date.weekday()]]
ws=wb["Saturday"]
wb.save("C:\\Users\\Farhan\\PycharmProjects\\Selenium_python\\folder\\date.xlsx")


class Keywords:

    def __init__(self, locator, row, weekday):

        self.locator = locator
        self.row = row
        self.weekday = weekday

    def Go_to_google(self):

        driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(self.locator)

    def Find_min_max_option(self):

        max = 0
        min = 10000
        for item in wait.until(EC.visibility_of_any_elements_located((By.XPATH, "//div[@class='wM6W7d']"))):
            if (len(item.text) > max):
                max = len(item.text)
            if (min > len(item.text)):
                min = len(item.text)
        for item in wait.until(EC.visibility_of_any_elements_located((By.XPATH, "//div[@class='wM6W7d']"))):
            if (len(item.text) == max):
                max_string = item.text
            if (len(item.text) == min):
                min_string = item.text
        ws_week=wb[self.weekday]
        ws_week['D'+str(self.row)]=max_string
        ws_week['E'+str(self.row)]=min_string
        wb.save("C:\\Users\\Farhan\\PycharmProjects\\Selenium_python\\folder\\date.xlsx")
keyword1 = Keywords(ws['C3'].value,3,calendar.day_name[my_date.weekday()])
keyword1.Go_to_google()
keyword1.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword2= Keywords(ws['C4'].value,4,calendar.day_name[my_date.weekday()])
keyword2.Go_to_google()
keyword2.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword3=Keywords(ws['C5'].value,5,calendar.day_name[my_date.weekday()])
keyword3.Go_to_google()
keyword3.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword4=Keywords(ws['C6'].value,6,calendar.day_name[my_date.weekday()])
keyword4.Go_to_google()
keyword4.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword5=Keywords(ws['C7'].value,7,calendar.day_name[my_date.weekday()])
keyword5.Go_to_google()
keyword5.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword6=Keywords(ws['C8'].value,8,calendar.day_name[my_date.weekday()])
keyword6.Go_to_google()
keyword6.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword7=Keywords(ws['C9'].value,9,calendar.day_name[my_date.weekday()])
keyword7.Go_to_google()
keyword7.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword8=Keywords(ws['C10'].value,10,calendar.day_name[my_date.weekday()])
keyword8.Go_to_google()
keyword8.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword9=Keywords(ws['C11'].value,11,calendar.day_name[my_date.weekday()])
keyword9.Go_to_google()
keyword9.Find_min_max_option()

driver.find_element(By.XPATH, '//input[@type="text"]').send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)

keyword10=Keywords(ws['C12'].value,12,calendar.day_name[my_date.weekday()])
keyword10.Go_to_google()
keyword10.Find_min_max_option()


driver.close()
